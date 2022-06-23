try:
    import os
    import re
    import warnings
    import datetime
    import traceback

    # import win32com.client
    import pandas as pd
    import numpy as np
    import xlwings as xw
    import openpyxl

    # import win32com.client
    pd.options.mode.chained_assignment = None  # default='warn'
except Exception as e:
    print(e)
    input('')


def exit_with_message(text):
    print(text)
    input()
    exit(0)


def excel_column_number(name):
    """Excel-style column name to number, e.g., A = 1, Z = 26, AA = 27, AAA = 703."""
    if isinstance(name, int):
        return name
    if name.isdigit():
        return int(name)
    n = 0
    for c in name:
        n = n * 26 + 1 + ord(c) - ord('A')
    return n - 1


def excel_column_name(n):
    """Number to Excel-style column name, e.g., 1 = A, 26 = Z, 27 = AA, 703 = AAA."""
    name = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        name = chr(r + ord('A')) + name
    return name


def check_workspace():
    items = os.listdir(WORK_DIR)
    if 'source_files' not in items:
        os.mkdir(WORK_DIR+r'\source_files')
        exit_with_message('Place files in source_files folder and rerun script')
    if 'script_files' not in items:
        os.mkdir(WORK_DIR+r'script_files')
    script_files = os.listdir(os.path.join(WORK_DIR, 'script_files'))
    file_path = os.path.join(WORK_DIR, 'script_files', '_last_info_update.csv')
    if '_last_info_update.csv' not in script_files:
        df = pd.DataFrame(columns=['filename', 'update_timestamp'])
        df.to_csv(file_path, index=False)
    else:
        df = pd.read_csv(file_path)
    return df


def load_filter_settings():
    wb = openpyxl.load_workbook(MAIN_FILE,
                                read_only=True, data_only=True, keep_links=False, keep_vba=False)
    ws = wb['welders statistic']
    d = {
        'joints_type': ws['B1'].value,
        'material': ws['D1'].value,
        'range_type': ws['B2'].value,
        'week': ws['B3'].value,
        'from_date': ws['B4'].value.date(),
        'to_date': ws['B5'].value.date(),
    }
    wb.close()
    return d


def load_generator_settings():
    split_text = ['JOINTS REPORT DATA', 'PROJECT DATA', 'LOG DATA', 'WELDERS DATA']
    d = pd.read_excel(MAIN_FILE, 'generator settings', header=None)
    d = d.values.tolist()
    for i, row in enumerate(d):
        d[i] = [n for n in row if isinstance(n, str)]
    data = [
        # JOINTS REPORT DATA
        [d[0][0], d[2][0], d[2][1], {d[1][i]: d[2][i] for i in range(2, len(d[1]))}],
        [d[0][0], d[3][0], d[3][1], {d[1][i]: d[3][i] for i in range(2, len(d[1]))}],
        [d[0][0], d[4][0], d[4][1], {d[1][i]: d[4][i] for i in range(2, len(d[1]))}],
        [d[0][0], d[5][0], d[5][1], {d[1][i]: d[5][i] for i in range(2, len(d[1]))}],
        # PROJECT DATA
        [d[6][0], d[8][0], d[8][1], {d[7][i]: d[8][i] for i in range(2, len(d[7]))}],
        [d[6][0], d[9][0], d[9][1], {d[7][i]: d[9][i] for i in range(2, len(d[7]))}],
        # LOG DATA
        [d[10][0], d[12][0], d[12][1], {d[11][i]: d[12][i] for i in range(2, len(d[11]))}],
        # WELDERS DATA
        [d[13][0], d[15][0], d[15][1], {d[14][i]: d[15][i] for i in range(2, len(d[14]))}]
    ]
    # column_names = []
    # for i in data:
    #     column_names.extend(list(i[-1].keys()))
    # unique_column_names = []
    # for i in column_names:
    #     if i not in unique_column_names:
    #         unique_column_names.append(i)
    return data


def is_last_update_equals(filename, last_update_df: pd.DataFrame):
    if last_update_df.loc[last_update_df['filename'] == filename].empty:
        last_update_df = last_update_df.append({'filename': filename, 'update_timestamp': 0}, ignore_index=True)
    modify_timestamp = round(os.path.getctime(filename))
    last_update_timestamp = round(last_update_df.loc[last_update_df['filename'] == filename].update_timestamp.values[0])
    if modify_timestamp != last_update_timestamp:
        last_update_df.loc[last_update_df['filename'] == filename, 'update_timestamp'] = modify_timestamp
        return False, last_update_df
    return True, last_update_df


def verify_files(file_names):
    file_names = [i[1] for i in file_names]
    files_in_folder = os.listdir(WORK_DIR+r'source_files')
    for file_in_folder in files_in_folder:
        if file_in_folder not in file_names and "~" not in file_in_folder:
            print(f'    WARNING! File {file_in_folder} did not listed in settings, but located in source_files')
    for filename in file_names:
        if not os.path.isfile(os.path.join(WORK_DIR, 'source_files', filename)):
            exit_with_message(f'    WARNING! {filename} didn\'t exist in source_files folder. Check it and restart')


def load_excel_data(file):
    type_, wb_name, sheet = file
    wb_name = WORK_DIR + rf'\{wb_name}'
    extraction = {
        'PROJECT DATA': {
            'header_row': 0,
            'skip_rows': [1],
            'columns': None
            # 'columns': "F,Q,BJ"
        },
        'JOINTS REPORT DATA': {
            'header_row': 0,
            'skip_rows': None,
            'columns': None
            # 'columns': "C,N,O,T:V"
        },
        'LOG DATA': {
            'header_row': 0,
            'skip_rows': None,
            'columns': None
            # 'columns': "H,J,X:AC,AK"
        },
        'WELDERS DATA': {
            'header_row': 0,
            'skip_rows': [1, 2],
            'columns': "A:S"
        }
    }
    warnings.simplefilter('ignore')
    engine = 'pyxlsb' if wb_name.lower().endswith('xlsb') else 'openpyxl'
    df = pd.read_excel(wb_name, sheet,
                       engine=engine,
                       skiprows=extraction[type_]['skip_rows'],
                       usecols=extraction[type_]['columns'])
    warnings.simplefilter('default')
    df = df.rename(columns={i: i.replace('\n', ' ') for i in df.columns.values.tolist()})
    # df.dropna(how='all', inplace=True)
    # df.dropna(axis=1, how='all', inplace=True)
    wb_name = wb_name.replace('source_files', 'script_files')
    df.to_csv(wb_name + '.csv', index=False)


def join_df_cols(df, col_id):
    if ',' in col_id:
        delimiter = ''
        col_alphas = col_id.split(',')
    else:
        delimiter = re.findall(r'\(.*?\)', col_id)[0]
        delimiter = re.sub(r'[()]', '', delimiter)
        col_alphas = re.split(r'\(.*?\)', col_id)
    return df.iloc[:, [excel_column_number(i) for i in col_alphas]].apply(lambda x: delimiter.join(str(x)), axis=1)

def add_rows_to_main_df(main_df, df_a, cols_to_add, criteria_col=None, create_col=None):
    df_a = df_a.fillna("")
    if criteria_col:
        criteria_col = excel_column_number(criteria_col)
    if create_col:
        df_a[create_col[0]] = create_col[1]
        cols_to_add[create_col[0]] = len(df_a.columns.values) - 1
    for col_name, col_id in cols_to_add.items():
        if ',' in str(col_id) or '(' in str(col_id):
            df_a[col_name] = join_df_cols(df_a, col_id)
            cols_to_add[col_name] = len(df_a.columns.values) - 1
    if criteria_col:
        df_a = df_a[df_a.apply(lambda x: 'ACCEPTED' in str(x.iloc[criteria_col]).upper(), axis=1)]
    indxs_to_add = [excel_column_number(i) for i in cols_to_add.values()]
    df_a = df_a.rename(
        {df_a.columns.values[indxs_to_add[i]]: list(cols_to_add.keys())[i] for i in range(len(indxs_to_add))}, axis=1)
    main_df = main_df.append(df_a.iloc[:, indxs_to_add], ignore_index=True)
    return main_df


def filter_welders(w1, w2, w3, welder_id):
    w1 = change_letters(w1, 'ru->en') if len(str(w1)) in range(4, 6) else np.nan
    w2 = change_letters(w2, 'ru->en') if len(str(w2)) in range(4, 6) else np.nan
    w3 = change_letters(w3, 'ru->en') if len(str(w3)) in range(4, 6) else np.nan
    if welder_id == 1:
        if pd.isna(w1):
            if not pd.isna(w2):
                return w2
            elif not pd.isna(w3):
                return w3
            return 'Unknown'
        return w1
    elif welder_id == 2:
        if w1 == w2:
            return np.nan
        if pd.isna(w2) and not pd.isna(w3):
            if w1 != w3:
                return w3
        return w2
    elif welder_id == 3:
        if w3 == w2 or w3 == w1:
            return np.nan
        return w3


def change_letters(text, mode):
    ru = 'ÐÐ’Ð•ÐšÐœÐžÐÐ¡Ð¢Ð£Ð¥'
    en = 'ABEKMOPCTYX'
    if mode == 'ru->en':
        return ''.join([en[ru.index(i)] if i in ru else i for i in text])
    elif mode == 'en->ru':
        return ''.join([ru[en.index(i)] if i in en else i for i in text])


def filter_material(material):
    if pd.isna(material):
        return 'Unknown'
    material = re.sub(r'[ \n]', '', material)
    material = material.upper()
    if '/' in material:
        material = material.split('/')
        material[0] = change_letters(material[0], 'ru->en')
        material[1] = change_letters(material[1], 'en->ru')
        material = '/'.join(material)
    return material


def repair_date(date):
    if pd.isna(date):
        return date
    if (
            str(date).isdigit() or
            isinstance(date, (float, int))
    ):
        return datetime.date(1900, 1, 1) + datetime.timedelta(days=int(date))
    return np.nan

def date_from_str(date, template):
    if pd.isna(date):
        return datetime.date(1990, 1, 1)
    if len(str(date)) > 11:
        template = template + ' %H:%M:%S'       
    return datetime.datetime.strptime(date, template).date()

def clean_trash(value):
    if isinstance(value, str):
        value = value.strip()
        value = value.replace('\n', ' ')
        value = ' '.join(value.split())
        if value in ['!', '0', 'UT', 'RT'] or not value:
            return np.nan
    elif isinstance(value, int):
        if value == 0:
            return np.nan
    return value


def filter_data(df):
    df['welding date'] = df['welding date'].apply(date_from_str, args=['%Y-%m-%d'])
    df['Material'] = df['Material'].apply(filter_material)
    df['actual control date'] = df['actual control date'].apply(repair_date)
    df['date RT'] = df['date RT'].apply(repair_date)
    df['date UT'] = df['date UT'].apply(repair_date)
    df['WELDER ID 1'] = df.apply(lambda x: filter_welders(x['WELDER ID 1'], x['WELDER ID 2'], x['WELDER ID 3'], 1),
                                 axis=1)
    df['WELDER ID 2'] = df.apply(lambda x: filter_welders(x['WELDER ID 1'], x['WELDER ID 2'], x['WELDER ID 3'], 2),
                                 axis=1)
    df['WELDER ID 3'] = df.apply(lambda x: filter_welders(x['WELDER ID 1'], x['WELDER ID 2'], x['WELDER ID 3'], 3),
                                 axis=1)
    for col in df.columns.values.tolist():
        df[col] = df[col].apply(clean_trash)
    return df


def create_defect_cols(df):
    d_cols = ['Single Porosity (Aa)',
              'Aligned Porosity (Ab)',
              'Cluster Porosity (Ac)',
              'Single slag inclusion (Ba)',
              'Aligned Slag Inclusion (Bb)',
              'Cluster Slag Inclusion (Bc)',
              'Single tungsten inclusion (Ca)',
              'Aligned Tungsten Inclusion (Cb)',
              'Cluster Tungsten Inclusion (Cc)',
              'Incomplete root penetration (Da)',
              'Incomplete Inter-Pass Fusion (Db)',
              'Incomplete bevel fusion (Dc)',
              'Longitudinal crack (Ea)',
              'Transverse crack (Eb)',
              'Branched crack (Ec)',
              'Root Concavity (Fa)',
              'Root Convexity (Fb)',
              'Undercut (Fc)',
              'High low bevel (Fd)',
              'Oxide Inclusion (O)']
    for d_col in d_cols:
        df[d_col] = df['DEFECT'].apply(count_text_entries, args=[re.findall(r'(?<=\().*?(?=\))', d_col)[0]])
    return df


def count_text_entries(base, text):
    n = 0
    if isinstance(base, str):
        base = change_letters(base.upper(), 'ru->en')
        n = len(
            re.findall(r'(?:\b|\d)'+text.upper()+r'(?:\b|\d)', base)
        )
    return n


def create_phase_col(df):
    df['PHASE'] = df['ISO'].apply(extract_phase)
    return df


def extract_phase(iso):
    match = re.findall(r'\b\d-\d-\d-', iso)
    if match:
        return 'Ph' + match[0].split('-')[1]
    return 'Unknown'


def create_shres_col(df):
    df['SHRES'] = df.apply(lambda x: extract_control_result(x['status RT'], x['status UT']), axis=1)
    return df


def extract_control_result(status_rt, status_ut):
    if not pd.isna(status_rt):
        if 'Ð³Ð¾Ð´ÐµÐ½' in status_rt.lower():
            return 'Acc'
        return 'Rej'
    if not pd.isna(status_ut):
        if 'Ð³Ð¾Ð´ÐµÐ½' in status_ut.lower():
            return 'Acc'
        return 'Rej'
    return 'No Control'


def order_columns(df):
    order = ['ISO+JOINT', 'ISO', 'SPOOL 1', 'F/S', 'WJ No', 'WJ TYPE', 'SCH', 'F/N',
             'WELDING VT RESULT', 'welding date', 'WELDER ID 1', 'WELDER ID 2', 'WELDER ID 3',
             'Area', 'PHASE', 'Material', 'WELDING method', 'INCH', 'THICKNESS', 'SHRES',
             'report RT', 'status RT', 'date RT', 'report UT', 'status UT', 'date UT', 'actual control date',

             'DEFECT', 'Single Porosity (Aa)', 'Aligned Porosity (Ab)', 'Cluster Porosity (Ac)',
             'Single slag inclusion (Ba)', 'Aligned Slag Inclusion (Bb)', 'Cluster Slag Inclusion (Bc)',
             'Single tungsten inclusion (Ca)', 'Aligned Tungsten Inclusion (Cb)', 'Cluster Tungsten Inclusion (Cc)',
             'Incomplete root penetration (Da)', 'Incomplete Inter-Pass Fusion (Db)', 'Incomplete bevel fusion (Dc)',
             'Longitudinal crack (Ea)', 'Transverse crack (Eb)', 'Branched crack (Ec)', 'Root Concavity (Fa)',
             'Root Convexity (Fb)', 'Undercut (Fc)', 'High low bevel (Fd)', 'Oxide Inclusion (O)']
    df = df[order]
    return df


def count_welder_statistic(df):
    f_settings = load_filter_settings()
    if f_settings['joints_type'] != 'ALL':
        if f_settings['joints_type'] == 'NEW JOINTS':
            filter_0 = df['F/N'] == 'F'
        else:
            filter_0 = df['F/N'] == 'N'
    else:
        filter_0 = df['F/N'].isin(['F', 'N'])
    if f_settings['material'] != 'ALL':
        filter_1 = df['Material'] == f_settings['material']
    else:
        filter_1 = ~pd.isna(df['Material'])
    if f_settings['range_type'] == 'WEEK':
        df['WEEK'] = df['welding date'].apply(lambda x: x.strftime("%V"))
        filter_2 = df['WEEK'] == str(f_settings['week'])
    elif f_settings['range_type'] == 'DATES':
        filter_2 = (f_settings['from_date'] <= df['welding date']) & (df['welding date'] <= f_settings['to_date'])
    else:
        filter_2 = df['F/N'].isin(['F', 'N'])
    filtered_df = df[filter_0 & filter_1 & filter_2]

    welders_dict = {}
    defect_cols = filtered_df.columns.values.tolist()[28:]
    for i in ['1', '2', '3']:
        stat_df = pd.DataFrame()
        temp_df = filtered_df[filtered_df[f'WELDER ID {i}'].notna()]
        stat_df['NAKS ID'] = temp_df[f'WELDER ID {i}'].unique()
        stat_df = stat_df.set_index('NAKS ID')
        stat_df['Q-ty welds'] = temp_df[f'WELDER ID {i}'].value_counts()
        stat_df['WELD Dinch'] = temp_df.groupby([f'WELDER ID {i}'])['INCH'].sum()
        temp_df.loc[:, 'controlled'] = temp_df.loc[:, 'SHRES'].isin(['Acc', 'Rej'])
        stat_df['TOTAL NDT'] = temp_df.groupby([f'WELDER ID {i}'])['controlled'].sum()
        temp_df.loc[:, 'acc'] = temp_df.loc[:, 'SHRES'].isin(['Acc'])
        stat_df['ACCEPTED'] = temp_df.groupby([f'WELDER ID {i}'])['acc'].sum()
        temp_df.loc[:, 'rej'] = temp_df.loc[:, 'SHRES'].isin(['Rej'])
        stat_df['REJECTED'] = temp_df.groupby([f'WELDER ID {i}'])['rej'].sum()
        g = temp_df.groupby([f'WELDER ID {i}'])
        stat_df['REJECTED INCH'] = g.apply(lambda x: x[x['rej'] == True]['INCH'].sum())
        stat_df['TOTAL NDT(up to 4 inch)'] = g.apply(lambda x: x[x['INCH'] <= 4]['controlled'].sum())
        stat_df['REJECTED(up to 4 inch)'] = g.apply(lambda x: x[x['INCH'] <= 4]['rej'].sum())
        stat_df['TOTAL NDT RESULT'] = 0
        stat_df[defect_cols] = temp_df.groupby([f'WELDER ID {i}'])[defect_cols].sum()
        welders_dict[f'WELDER ID {i}'] = stat_df
        del temp_df

    stat_df = pd.DataFrame()
    for v in welders_dict.values():
        stat_df = stat_df.add(v, fill_value=0)
    stat_df['Repair VS Welds'] = (stat_df['REJECTED'] / stat_df['Q-ty welds']).apply(round, args=[1])
    stat_df['Repair VS NDT'] = (stat_df['REJECTED'] / stat_df['TOTAL NDT']).apply(round, args=[1])
    stat_df['Repair VS NDT(up to 4 inch)'] = (stat_df['REJECTED(up to 4 inch)'] / stat_df['TOTAL NDT(up to 4 inch)']).apply(round, args=[1])

    col_order = ['Q-ty welds', 'WELD Dinch', 'TOTAL NDT', 'ACCEPTED', 'REJECTED', 'REJECTED INCH', 'Repair VS Welds',
                 'Repair VS NDT', 'TOTAL NDT(up to 4 inch)', 'REJECTED(up to 4 inch)', 'Repair VS NDT(up to 4 inch)',
                 'TOTAL NDT RESULT', *defect_cols]
    stat_df = stat_df[col_order]
    stat_df = stat_df.fillna(0)
    return stat_df


def is_all_ids_in_excel(excel_ids, data_ids):
    for id_ in data_ids:
        if id_ not in excel_ids:
            print(f"    Warning! {id_} miss in welders performance table")


def wright_welder_statistic_to_excel(df):
    wb = xw.Book('_Welder Performance.xlsm')
    ws = wb.sheets['welders statistic']
    naks_ids_raw = ws.range('E8').expand('down').value
    naks_ids = []
    for naks_id in naks_ids_raw:
        if isinstance(naks_id, float):
            naks_id = int(naks_id)
        naks_ids.append(str(naks_id))
    rec_array = df.to_records().tolist()
    rec_array = {i[0]: i[1:] for i in rec_array}
    is_all_ids_in_excel(naks_ids, list(rec_array.keys()))
    ordered_rec_aray = []
    for naks in naks_ids:
        if rec_array.get(naks):
            ordered_rec_aray.append(rec_array[naks])
        else:
            ordered_rec_aray.append([0] * 32)
    ws.range('F8').value = ordered_rec_aray


def wright_ndt_result_to_excel(df):
    wb = xw.Book('_Welder Performance.xlsm')
    ws = wb.sheets['NDT result']
    df = df.fillna('')
    rec_array = df.to_records().tolist()
    for i in range(len(rec_array)):
        rec_array[i] = list(rec_array[i])
        for i2 in range(len(rec_array[i])):
            if isinstance(rec_array[i][i2], datetime.date):
                t = rec_array[i][i2]
                rec_array[i][i2] = datetime.datetime(t.year, t.month, t.day)
    ws.range('A2').value = rec_array


def main():
    print('[-] Welder statistic generator start work')
    last_update_info = check_workspace()
    files = load_generator_settings()
    verify_files(files)
    joints_report = pd.DataFrame()
    project = pd.DataFrame()
    log = pd.DataFrame()
    welders = pd.DataFrame()
    files_to_update = []
    print('[-] Check if files updated')
    for file_ in files:
        file_[1] = os.path.join('source_files', file_[1])
        result, last_update_info = is_last_update_equals(WORK_DIR + fr'\{file_[1]}', last_update_info)
        if not result:
            print(f"    Upload data from {file_[1]}")
            files_to_update.append(file_)
            load_excel_data(file_[0:3])
    if files_to_update:
        print(f'[-] Processing updated files')
        for file_ in files:
            print(f'    [{file_[0]}]', file_[1])
            file_[1] = file_[1].replace('source_files', 'script_files')
            df_temp = pd.read_csv(WORK_DIR + fr'\{file_[1]}' + '.csv', low_memory=False)
            if file_[0] == 'JOINTS REPORT DATA':
                joints_report = add_rows_to_main_df(
                    main_df=joints_report,
                    df_a=df_temp,
                    cols_to_add=file_[3],
                    criteria_col=file_[3]['WELDING VT RESULT'],
                    create_col=['F/N', 'F' if 'NEW' in file_[1] else 'N']
                )
            elif file_[0] == 'PROJECT DATA':
                project = add_rows_to_main_df(
                    main_df=project,
                    df_a=df_temp,
                    cols_to_add=file_[3]
                )
            elif file_[0] == 'LOG DATA':
                log = add_rows_to_main_df(
                    main_df=log,
                    df_a=df_temp,
                    cols_to_add=file_[3]
                )
            elif file_[0] == 'WELDERS DATA':
                welders = add_rows_to_main_df(
                    main_df=welders,
                    df_a=df_temp,
                    cols_to_add=file_[3]
                )
        print('[-] Create NDT result table with updated data')
        ndt_result = pd.merge(joints_report, project, on='ISO+JOINT', how='left')
        ndt_result = ndt_result.merge(log, on='ISO+JOINT', how='left')
        ndt_result = filter_data(ndt_result)
        ndt_result = create_phase_col(ndt_result)
        ndt_result = create_defect_cols(ndt_result)
        ndt_result = create_shres_col(ndt_result)
        ndt_result = order_columns(ndt_result)
        ndt_result.to_csv(os.path.join(WORK_DIR, 'script_files', 'ndt_result.csv'), index=False)
        wright_ndt_result_to_excel(ndt_result)
        last_update_info.to_csv(os.path.join(WORK_DIR, 'script_files', '_last_info_update.csv'), index=False)
    else:
        print('[-] All files up-to-date. Load data from NDT result table')
        ndt_result = pd.read_csv(os.path.join(WORK_DIR, 'script_files', 'ndt_result.csv'), low_memory=False,
                                 parse_dates=['welding date'], date_parser=pd.to_datetime)
        ndt_result['welding date'] = ndt_result['welding date'].apply(lambda x: x.date())
    print('[-] Welder statistic processing')
    welder_statistic = count_welder_statistic(ndt_result)
    wright_welder_statistic_to_excel(welder_statistic)
    print('[-] Work done!')


if __name__ == '__main__':
    try:
        wb = xw.Book('_Welder Performance.xlsm')
        ws = wb.sheets['welders statistic']
        WORK_DIR = ws.range('D3').value
        MAIN_FILE = WORK_DIR + '_Welder Performance.xlsm'
        print(MAIN_FILE)
        print(os.getcwd())
        main()
    except Exception as e:
        print(traceback.format_exc())
        print(e)
    finally:
        exit_with_message('Press Enter to exit   SER-xWMY-bnRF-ZDon2 ')